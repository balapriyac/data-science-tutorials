"""
pivot_report.py
Generate a formatted multi-tab summary/pivot report from raw transaction data.

Dependencies: pandas, openpyxl, matplotlib
Install:      pip install pandas openpyxl matplotlib

Usage:
    python pivot_report.py --input transactions.xlsx --date-col "Date" --value-col "Amount" --group-cols "Category" "Region"
    python pivot_report.py --input sales.xlsx --date-col "Sale Date" --value-col "Revenue" --group-cols "Product" --top-n 10
    python pivot_report.py --input data.xlsx  --date-col "Date" --value-col "Amount" --group-cols "Category" --config report_config.json
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter

# ── CONFIG (override via CLI) ─────────────────────────────────────────────────
INPUT_FILE  = "transactions.xlsx"
OUTPUT_FILE = "pivot_report.xlsx"
DATE_COL    = "Date"
VALUE_COL   = "Amount"
GROUP_COLS  = ["Category"]      # One or more grouping columns
TOP_N       = 10                # Rows to show in top-N ranking
SHEET       = 0
# ─────────────────────────────────────────────────────────────────────────────

HEADER_BLUE  = "1F4E79"
ACCENT_GREEN = "70AD47"
ACCENT_RED   = "FF0000"
ALT_ROW      = "DEEAF1"


def style_ws(ws, header_color: str = HEADER_BLUE, alt_row: bool = True) -> None:
    """Apply consistent header + alternating row styling."""
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=header_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    if alt_row:
        alt_fill = PatternFill("solid", fgColor=ALT_ROW)
        for row_idx in range(2, ws.max_row + 1):
            if row_idx % 2 == 0:
                for cell in ws[row_idx]:
                    cell.fill = alt_fill

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.freeze_panes = "A2"


def add_conditional_color(ws, col_letter: str, min_row: int, max_row: int) -> None:
    """Manually color top value green, bottom value red in a column."""
    values = []
    for row in range(min_row, max_row + 1):
        val = ws[f"{col_letter}{row}"].value
        if isinstance(val, (int, float)):
            values.append((val, row))
    if not values:
        return
    values.sort()
    ws[f"{col_letter}{values[-1][1]}"].font  = Font(bold=True, color=ACCENT_GREEN)
    ws[f"{col_letter}{values[0][1]}"].font   = Font(bold=True, color=ACCENT_RED)


def build_report(input_file: str, output_file: str, date_col: str, value_col: str,
                 group_cols: list[str], top_n: int, sheet) -> None:
    src = Path(input_file)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    print(f"Reading: {src}")
    df = pd.read_excel(src, sheet_name=sheet)
    print(f"  {len(df):,} rows, {len(df.columns)} columns")

    # Validate columns
    for col in [date_col, value_col] + group_cols:
        if col not in df.columns:
            sys.exit(f"[ERROR] Column not found: '{col}'\n  Available: {list(df.columns)}")

    # Parse dates & values
    df[date_col]  = pd.to_datetime(df[date_col], infer_datetime_format=True, errors="coerce")
    df[value_col] = pd.to_numeric(df[value_col], errors="coerce")
    df.dropna(subset=[date_col, value_col], inplace=True)

    df["_Year"]  = df[date_col].dt.year
    df["_Month"] = df[date_col].dt.to_period("M").astype(str)

    out = Path(output_file)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:

        # ── Tab 1: Overview ──────────────────────────────────────────────────
        total      = df[value_col].sum()
        avg_txn    = df[value_col].mean()
        count      = len(df)
        date_range = f"{df[date_col].min().date()} → {df[date_col].max().date()}"

        overview = pd.DataFrame([
            {"Metric": "Total Value",       "Value": f"{total:,.2f}"},
            {"Metric": "Transaction Count", "Value": f"{count:,}"},
            {"Metric": "Average per Row",   "Value": f"{avg_txn:,.2f}"},
            {"Metric": "Date Range",        "Value": date_range},
            {"Metric": "Group Columns",     "Value": ", ".join(group_cols)},
            {"Metric": "Generated",         "Value": datetime.now().strftime("%Y-%m-%d %H:%M")},
        ])
        overview.to_excel(writer, sheet_name="Overview", index=False)

        # ── Tab 2: By Group ──────────────────────────────────────────────────
        group_summary = (
            df.groupby(group_cols)[value_col]
            .agg(Total="sum", Count="count", Average="mean", Min="min", Max="max")
            .reset_index()
            .sort_values("Total", ascending=False)
        )
        group_summary["% of Total"] = (group_summary["Total"] / total * 100).round(2)
        group_summary.to_excel(writer, sheet_name="By Group", index=False)

        # ── Tab 3: Monthly Trend ─────────────────────────────────────────────
        monthly = (
            df.groupby("_Month")[value_col]
            .agg(Total="sum", Count="count")
            .reset_index()
            .rename(columns={"_Month": "Month"})
        )
        monthly["MoM Change"] = monthly["Total"].pct_change().mul(100).round(2)
        monthly.to_excel(writer, sheet_name="Monthly Trend", index=False)

        # ── Tab 4: Top N ─────────────────────────────────────────────────────
        top_df = group_summary.head(top_n).copy()
        top_df.to_excel(writer, sheet_name=f"Top {top_n}", index=False)

        # ── Tab 5: Year-over-Year ────────────────────────────────────────────
        yoy = (
            df.groupby(["_Year"] + group_cols)[value_col]
            .sum()
            .unstack("_Year")
            .reset_index()
        )
        yoy.columns = [str(c) for c in yoy.columns]
        yoy.to_excel(writer, sheet_name="Year-over-Year", index=False)

    # ── Post-processing: styling + charts ────────────────────────────────────
    wb = load_workbook(out)

    for sname in wb.sheetnames:
        style_ws(wb[sname])

    # Color top/bottom in By Group
    ws_group = wb["By Group"]
    total_col_letter = "D"  # Total column (0-indexed: group_cols + Total=D)
    if len(group_cols) == 1:
        total_col_letter = "B"
    elif len(group_cols) == 2:
        total_col_letter = "C"
    add_conditional_color(ws_group, total_col_letter, 2, ws_group.max_row)

    # Bar chart — By Group (top 15)
    ws_group = wb["By Group"]
    n_rows   = min(ws_group.max_row, 16)   # header + up to 15 data rows
    bar = BarChart()
    bar.type   = "col"
    bar.title  = f"Total {value_col} by {', '.join(group_cols)}"
    bar.y_axis.title = value_col
    bar.x_axis.title = ", ".join(group_cols)
    bar.width  = 24
    bar.height = 14

    # Total column index: len(group_cols) + 1 (1-based)
    total_col_idx = len(group_cols) + 1
    data   = Reference(ws_group, min_col=total_col_idx, min_row=1, max_row=n_rows)
    cats   = Reference(ws_group, min_col=1, min_row=2,  max_row=n_rows)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws_group.add_chart(bar, f"A{n_rows + 3}")

    # Line chart — Monthly Trend
    ws_monthly = wb["Monthly Trend"]
    n_months   = ws_monthly.max_row
    line = LineChart()
    line.title  = f"Monthly {value_col} Trend"
    line.y_axis.title = value_col
    line.x_axis.title = "Month"
    line.width  = 24
    line.height = 14
    data_ref = Reference(ws_monthly, min_col=2, min_row=1, max_row=n_months)
    cats_ref = Reference(ws_monthly, min_col=1, min_row=2, max_row=n_months)
    line.add_data(data_ref, titles_from_data=True)
    line.set_categories(cats_ref)
    ws_monthly.add_chart(line, f"A{n_months + 3}")

    wb.save(out)
    print(f"\nReport written to: {out.resolve()}")
    print("  Tabs: Overview | By Group | Monthly Trend | Top N | Year-over-Year")
    print("  Charts embedded in 'By Group' and 'Monthly Trend' tabs")


def main():
    parser = argparse.ArgumentParser(description="Generate pivot summary report from Excel data.")
    parser.add_argument("--input",      default=INPUT_FILE)
    parser.add_argument("--output",     default=OUTPUT_FILE)
    parser.add_argument("--date-col",   default=DATE_COL,   help="Date column name")
    parser.add_argument("--value-col",  default=VALUE_COL,  help="Numeric value column name")
    parser.add_argument("--group-cols", nargs="+", default=GROUP_COLS, help="Grouping column(s)")
    parser.add_argument("--top-n",      type=int, default=TOP_N, help="Rows in Top N tab")
    parser.add_argument("--sheet",      default=SHEET)
    args = parser.parse_args()

    build_report(
        input_file=args.input,
        output_file=args.output,
        date_col=args.date_col,
        value_col=args.value_col,
        group_cols=args.group_cols,
        top_n=args.top_n,
        sheet=args.sheet,
    )


if __name__ == "__main__":
    main()

