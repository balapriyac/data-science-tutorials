"""
data_cleaner.py
Clean and standardize messy Excel/CSV exports using configurable rules.

Dependencies: pandas, openpyxl
Install:      pip install pandas openpyxl

Usage:
    python data_cleaner.py --input export.xlsx --output cleaned.xlsx
    python data_cleaner.py --input export.csv  --output cleaned.xlsx --config cleaning_rules.json

Example cleaning_rules.json:
{
    "strip_whitespace":  ["Name", "Email", "Address"],
    "title_case":        ["Name", "City"],
    "upper_case":        ["State", "Country Code"],
    "lower_case":        ["Email"],
    "date_format":       {"columns": ["Date", "DOB"], "output_format": "%Y-%m-%d"},
    "phone_normalize":   ["Phone", "Mobile"],
    "remove_blank_rows": true,
    "remove_duplicates": ["Email"]
}
"""

import argparse
import json
import re
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Side, Border
from openpyxl.utils import get_column_letter

# ── DEFAULT CONFIG (used when no --config file is provided) ──────────────────
DEFAULT_RULES = {
    "strip_whitespace":  [],            # Column names to strip
    "title_case":        [],
    "upper_case":        [],
    "lower_case":        [],
    "date_format":       {},            # {"columns": [...], "output_format": "%Y-%m-%d"}
    "phone_normalize":   [],
    "remove_blank_rows": True,
    "remove_duplicates": [],            # Columns to use as dedup key
}
# ─────────────────────────────────────────────────────────────────────────────

DATE_FORMATS = [
    "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y",
    "%d %b %Y", "%d %B %Y", "%Y%m%d", "%b %d, %Y",
]


def parse_date(val: str, output_fmt: str) -> tuple[str, bool]:
    """Try known date formats; return (result, success)."""
    for fmt in DATE_FORMATS:
        try:
            from datetime import datetime
            return datetime.strptime(str(val).strip(), fmt).strftime(output_fmt), True
        except ValueError:
            continue
    return val, False


def normalize_phone(val: str) -> str:
    """Strip all non-digit characters, reformat to (XXX) XXX-XXXX if 10 digits."""
    digits = re.sub(r"\D", "", str(val))
    if len(digits) == 10:
        return f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
    if len(digits) == 11 and digits[0] == "1":
        return f"+1 ({digits[1:4]}) {digits[4:7]}-{digits[7:]}"
    return val  # Return unchanged if pattern doesn't match


def clean_dataframe(df: pd.DataFrame, rules: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Apply cleaning rules. Returns (cleaned_df, changelog_df).
    Changelog has columns: Row, Column, Original, Cleaned, Rule.
    """
    changes = []
    df = df.copy()

    def record(row_idx, col, original, cleaned, rule):
        if str(original) != str(cleaned):
            changes.append({
                "Row":      row_idx + 2,   # +2: 1-based + header row
                "Column":   col,
                "Original": original,
                "Cleaned":  cleaned,
                "Rule":     rule,
            })

    # Remove blank rows
    if rules.get("remove_blank_rows"):
        before = len(df)
        df.dropna(how="all", inplace=True)
        df = df[df.apply(lambda r: r.astype(str).str.strip().ne("").any(), axis=1)]
        removed = before - len(df)
        if removed:
            print(f"  Removed {removed} blank row(s)")

    # Strip whitespace
    for col in rules.get("strip_whitespace", []):
        if col not in df.columns:
            continue
        for i, val in enumerate(df[col]):
            cleaned = str(val).strip() if pd.notna(val) else val
            record(i, col, val, cleaned, "strip_whitespace")
            df.at[df.index[i], col] = cleaned

    # Case transforms
    for col in rules.get("title_case", []):
        if col not in df.columns:
            continue
        for i, val in enumerate(df[col]):
            if pd.notna(val):
                cleaned = str(val).strip().title()
                record(i, col, val, cleaned, "title_case")
                df.at[df.index[i], col] = cleaned

    for col in rules.get("upper_case", []):
        if col not in df.columns:
            continue
        for i, val in enumerate(df[col]):
            if pd.notna(val):
                cleaned = str(val).strip().upper()
                record(i, col, val, cleaned, "upper_case")
                df.at[df.index[i], col] = cleaned

    for col in rules.get("lower_case", []):
        if col not in df.columns:
            continue
        for i, val in enumerate(df[col]):
            if pd.notna(val):
                cleaned = str(val).strip().lower()
                record(i, col, val, cleaned, "lower_case")
                df.at[df.index[i], col] = cleaned

    # Date normalization
    date_cfg = rules.get("date_format", {})
    if date_cfg:
        out_fmt = date_cfg.get("output_format", "%Y-%m-%d")
        for col in date_cfg.get("columns", []):
            if col not in df.columns:
                continue
            error_col = f"{col}_parse_error"
            df[error_col] = ""
            for i, val in enumerate(df[col]):
                if pd.isna(val) or str(val).strip() == "":
                    continue
                cleaned, ok = parse_date(val, out_fmt)
                if ok:
                    record(i, col, val, cleaned, "date_format")
                    df.at[df.index[i], col] = cleaned
                else:
                    df.at[df.index[i], error_col] = "PARSE_FAILED"

    # Phone normalization
    for col in rules.get("phone_normalize", []):
        if col not in df.columns:
            continue
        for i, val in enumerate(df[col]):
            if pd.notna(val) and str(val).strip():
                cleaned = normalize_phone(str(val))
                record(i, col, val, cleaned, "phone_normalize")
                df.at[df.index[i], col] = cleaned

    # Deduplication
    dedup_cols = rules.get("remove_duplicates", [])
    if dedup_cols:
        valid_cols = [c for c in dedup_cols if c in df.columns]
        if valid_cols:
            before = len(df)
            df.drop_duplicates(subset=valid_cols, keep="first", inplace=True)
            removed = before - len(df)
            if removed:
                print(f"  Removed {removed} duplicate row(s) on {valid_cols}")

    df.reset_index(drop=True, inplace=True)
    changelog = pd.DataFrame(changes)
    return df, changelog


def style_sheets(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    change_fill = PatternFill("solid", fgColor="FFF2CC")

    for sname in wb.sheetnames:
        ws = wb[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        if sname == "Change Log":
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = change_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.freeze_panes = "A2"
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Clean and standardize Excel/CSV exports.")
    parser.add_argument("--input",  required=True,  help="Input file (.xlsx or .csv)")
    parser.add_argument("--output", default="cleaned.xlsx")
    parser.add_argument("--config", default=None,   help="JSON rules file (optional)")
    parser.add_argument("--sheet",  default=0,      help="Sheet name or index (default: 0)")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    rules = DEFAULT_RULES.copy()
    if args.config:
        cfg_path = Path(args.config)
        if not cfg_path.exists():
            sys.exit(f"[ERROR] Config file not found: {cfg_path}")
        with open(cfg_path) as fh:
            rules.update(json.load(fh))

    print(f"Reading: {src}")
    if src.suffix.lower() == ".csv":
        df = pd.read_csv(src, dtype=str)
    else:
        df = pd.read_excel(src, sheet_name=args.sheet, dtype=str)
    print(f"  {len(df):,} rows, {len(df.columns)} columns")

    print("\nApplying cleaning rules...")
    cleaned, changelog = clean_dataframe(df, rules)
    print(f"  {len(changelog):,} cell(s) modified")
    print(f"  {len(cleaned):,} rows in output")

    out = Path(args.output)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        cleaned.to_excel(writer, sheet_name="Cleaned Data", index=False)
        if not changelog.empty:
            changelog.to_excel(writer, sheet_name="Change Log", index=False)
        else:
            pd.DataFrame([{"Note": "No changes were made."}]).to_excel(
                writer, sheet_name="Change Log", index=False
            )

    style_sheets(out)
    print(f"\nOutput written to: {out.resolve()}")
    print("  'Cleaned Data' tab — cleaned output")
    print("  'Change Log' tab  — every cell that was modified")


if __name__ == "__main__":
    main()
  
