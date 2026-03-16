"""
sheet_splitter.py
Split one Excel sheet into separate files based on unique values in a column.

Dependencies: pandas, openpyxl
Install:      pip install pandas openpyxl

Usage:
    python sheet_splitter.py --input master.xlsx --split-col "Region"
    python sheet_splitter.py --input sales.xlsx  --split-col "Department" --output-dir ./splits
    python sheet_splitter.py --input report.xlsx --split-col "Manager" --name-template "Report_{value}_{date}.xlsx"
    python sheet_splitter.py --input data.xlsx   --split-col "Region" --email-map emails.csv

Example emails.csv:
    Value,Email
    North,north.manager@company.com
    South,south.manager@company.com
"""

import argparse
import smtplib
import sys
from datetime import datetime
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG (override via CLI flags) ──────────────────────────────────────────
INPUT_FILE    = "master.xlsx"
SPLIT_COLUMN  = "Region"
OUTPUT_DIR    = "./output_splits"
SHEET_NAME    = 0                   # Sheet index or name
NAME_TEMPLATE = "{value}.xlsx"      # Placeholders: {value}, {date}, {datetime}
EMAIL_MAP_CSV = None                # CSV with Value,Email columns (optional)

# SMTP settings (only needed if using --email-map)
SMTP_HOST     = "smtp.gmail.com"
SMTP_PORT     = 587
SMTP_USER     = ""                  # Your email address
SMTP_PASS     = ""                  # App password
EMAIL_SUBJECT = "Your report: {value}"
EMAIL_BODY    = "Hi,\n\nPlease find attached the report for {value}.\n\nRegards"
# ─────────────────────────────────────────────────────────────────────────────


def safe_filename(val: str) -> str:
    """Strip characters unsafe for filenames."""
    return "".join(c for c in str(val) if c.isalnum() or c in " _-").strip()


def build_filename(template: str, value: str) -> str:
    now = datetime.now()
    return template.format(
        value=safe_filename(value),
        date=now.strftime("%Y-%m-%d"),
        datetime=now.strftime("%Y%m%d_%H%M%S"),
    )


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    ws = wb.active
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = "A2"
    wb.save(path)


def send_email(smtp_cfg: dict, to_addr: str, subject: str,
               body: str, attachment_path: Path) -> None:
    msg = MIMEMultipart()
    msg["From"]    = smtp_cfg["user"]
    msg["To"]      = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain"))

    with open(attachment_path, "rb") as fh:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(fh.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition",
                    f'attachment; filename="{attachment_path.name}"')
    msg.attach(part)

    with smtplib.SMTP(smtp_cfg["host"], smtp_cfg["port"]) as server:
        server.ehlo()
        server.starttls()
        server.login(smtp_cfg["user"], smtp_cfg["pass"])
        server.sendmail(smtp_cfg["user"], to_addr, msg.as_string())


def split_file(input_file: str, split_col: str, output_dir: str,
               sheet, name_template: str, email_map_csv: str | None,
               smtp_cfg: dict) -> None:
    src = Path(input_file)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    print(f"Reading: {src}")
    df = pd.read_excel(src, sheet_name=sheet, dtype=str)
    print(f"  {len(df):,} rows, {len(df.columns)} columns")

    if split_col not in df.columns:
        sys.exit(f"[ERROR] Column '{split_col}' not found.\n"
                 f"  Available: {list(df.columns)}")

    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Load email map if provided
    email_map: dict[str, str] = {}
    if email_map_csv:
        em_path = Path(email_map_csv)
        if not em_path.exists():
            sys.exit(f"[ERROR] Email map not found: {em_path}")
        em_df = pd.read_csv(em_path, dtype=str)
        if not {"Value", "Email"}.issubset(em_df.columns):
            sys.exit("[ERROR] Email map CSV must have 'Value' and 'Email' columns.")
        email_map = dict(zip(em_df["Value"], em_df["Email"]))

    unique_values = df[split_col].dropna().unique()
    print(f"\nSplit column: '{split_col}'")
    print(f"Unique values: {len(unique_values)}\n")

    results = []

    for val in sorted(unique_values):
        subset = df[df[split_col] == val].reset_index(drop=True)
        filename = build_filename(name_template, val)
        out_path = out_dir / filename

        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            subset.to_excel(writer, sheet_name=str(val)[:31], index=False)

        style_workbook(out_path)

        email_status = "—"
        if email_map:
            to_addr = email_map.get(str(val))
            if to_addr:
                try:
                    send_email(
                        smtp_cfg,
                        to_addr,
                        EMAIL_SUBJECT.format(value=val),
                        EMAIL_BODY.format(value=val),
                        out_path,
                    )
                    email_status = f"Sent → {to_addr}"
                except Exception as e:
                    email_status = f"FAILED: {e}"
            else:
                email_status = "No email mapping"

        results.append({
            "Value":  val,
            "Rows":   len(subset),
            "File":   filename,
            "Email":  email_status,
        })
        print(f"  ✓ {str(val):30s} {len(subset):>5} rows  →  {filename}")
        if email_map:
            print(f"    Email: {email_status}")

    total_rows = sum(r["Rows"] for r in results)
    print(f"\nTotal: {len(results)} files, {total_rows:,} rows")
    print(f"Output directory: {out_dir.resolve()}")

    # Write manifest
    manifest_path = out_dir / "_manifest.csv"
    pd.DataFrame(results).to_csv(manifest_path, index=False)
    print(f"Manifest: {manifest_path.resolve()}")


def main():
    parser = argparse.ArgumentParser(description="Split Excel sheet into separate files by column value.")
    parser.add_argument("--input",         default=INPUT_FILE)
    parser.add_argument("--split-col",     default=SPLIT_COLUMN,  help="Column to split on")
    parser.add_argument("--output-dir",    default=OUTPUT_DIR)
    parser.add_argument("--sheet",         default=SHEET_NAME)
    parser.add_argument("--name-template", default=NAME_TEMPLATE,
                        help="Filename template. Use {value}, {date}, {datetime}")
    parser.add_argument("--email-map",     default=EMAIL_MAP_CSV,
                        help="CSV file with Value,Email columns")
    parser.add_argument("--smtp-host",     default=SMTP_HOST)
    parser.add_argument("--smtp-port",     type=int, default=SMTP_PORT)
    parser.add_argument("--smtp-user",     default=SMTP_USER)
    parser.add_argument("--smtp-pass",     default=SMTP_PASS)
    args = parser.parse_args()

    smtp_cfg = {
        "host": args.smtp_host,
        "port": args.smtp_port,
        "user": args.smtp_user,
        "pass": args.smtp_pass,
    }

    split_file(
        input_file=args.input,
        split_col=args.split_col,
        output_dir=args.output_dir,
        sheet=args.sheet,
        name_template=args.name_template,
        email_map_csv=args.email_map,
        smtp_cfg=smtp_cfg,
    )


if __name__ == "__main__":
    main()

