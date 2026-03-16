"""
duplicate_finder.py
Find exact and near-duplicate rows in an Excel file based on key columns.

Dependencies: pandas, openpyxl, rapidfuzz
Install:      pip install pandas openpyxl rapidfuzz

Usage:
    python duplicate_finder.py --input data.xlsx --key-cols "Email" "Full Name"
    python duplicate_finder.py --input data.xlsx --key-cols "Email" --fuzzy-threshold 90
    python duplicate_finder.py --input data.xlsx --key-cols "Name" "Phone" --output flagged.xlsx
"""

import argparse
import sys
from itertools import combinations
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz

# ── CONFIG (override via CLI flags) ──────────────────────────────────────────
INPUT_FILE       = "data.xlsx"          # Input Excel file
OUTPUT_FILE      = "duplicates_flagged.xlsx"
KEY_COLUMNS      = ["Email", "Name"]    # Columns to check for duplicates
FUZZY_THRESHOLD  = 85                   # 0–100; matches above this are flagged
SHEET_NAME       = 0                    # Sheet index or name; 0 = first sheet
# ─────────────────────────────────────────────────────────────────────────────

EXACT_COLOR  = "FF4500"   # Orange-red  — exact duplicates
FUZZY_COLOR  = "FFD700"   # Gold        — fuzzy/near duplicates
HEADER_COLOR = "1F4E79"


def normalize(val) -> str:
    """Lowercase, strip, collapse whitespace."""
    return " ".join(str(val).lower().split()) if pd.notna(val) else ""


def find_duplicates(df: pd.DataFrame, key_cols: list[str],
                    threshold: int) -> pd.DataFrame:
    """Return df annotated with _dup_group, _dup_type, _dup_score columns."""
    n = len(df)
    group_id   = [None] * n
    dup_type   = [""] * n
    dup_score  = [None] * n
    group_counter = 1

    # Build normalized composite key for each row
    keys = [
        " | ".join(normalize(df.at[i, c]) for c in key_cols if c in df.columns)
        for i in range(n)
    ]

    # Step 1 — exact duplicates
    seen: dict[str, int] = {}
    for i, key in enumerate(keys):
        if key in seen:
            j = seen[key]
            # Assign group
            if group_id[j] is None:
                group_id[j] = group_counter
                dup_type[j] = "Exact"
                dup_score[j] = 100
                group_counter += 1
            group_id[i] = group_id[j]
            dup_type[i] = "Exact"
            dup_score[i] = 100
        else:
            seen[key] = i

    # Step 2 — fuzzy duplicates among non-exact rows
    unmatched = [i for i in range(n) if group_id[i] is None]
    for i, j in combinations(unmatched, 2):
        score = fuzz.token_sort_ratio(keys[i], keys[j])
        if score >= threshold:
            # Assign shared group
            gi, gj = group_id[i], group_id[j]
            if gi is None and gj is None:
                group_id[i] = group_counter
                group_id[j] = group_counter
                dup_type[i] = dup_type[j] = "Fuzzy"
                dup_score[i] = dup_score[j] = score
                group_counter += 1
            elif gi is not None and gj is None:
                group_id[j] = gi
                dup_type[j] = "Fuzzy"
                dup_score[j] = score
            elif gj is not None and gi is None:
                group_id[i] = gj
                dup_type[i] = "Fuzzy"
                dup_score[i] = score
            else:
                # Update score if this pair has higher confidence
                if score > (dup_score[i] or 0):
                    dup_score[i] = score
                if score > (dup_score[j] or 0):
                    dup_score[j] = score

    df = df.copy()
    df.insert(0, "_dup_score",  dup_score)
    df.insert(0, "_dup_type",   dup_type)
    df.insert(0, "_dup_group",  group_id)
    return df


def style_output(path: Path, df_annotated: pd.DataFrame) -> None:
    wb = load_workbook(path)
    ws = wb["Flagged Data"]

    # Header style
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")

    # Color duplicate rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        dup_type_val = ws.cell(row=row_idx, column=3).value  # _dup_type col
        if dup_type_val == "Exact":
            fill = PatternFill("solid", fgColor=EXACT_COLOR)
        elif dup_type_val == "Fuzzy":
            fill = PatternFill("solid", fgColor=FUZZY_COLOR)
        else:
            continue
        for cell in row:
            cell.fill = fill

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20
    ws.freeze_panes = "A2"

    # Summary sheet
    ws_sum = wb["Summary"]
    for cell in ws_sum[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Find duplicate rows in Excel.")
    parser.add_argument("--input",           default=INPUT_FILE)
    parser.add_argument("--output",          default=OUTPUT_FILE)
    parser.add_argument("--key-cols",        nargs="+", default=KEY_COLUMNS,
                        help="Column names to use as duplicate key")
    parser.add_argument("--fuzzy-threshold", type=int, default=FUZZY_THRESHOLD,
                        help="Fuzzy match score (0-100) to flag near-duplicates")
    parser.add_argument("--sheet",           default=SHEET_NAME)
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        sys.exit(f"[ERROR] File not found: {path}")

    print(f"Reading: {path}")
    df = pd.read_excel(path, sheet_name=args.sheet, dtype=str)
    print(f"  {len(df):,} rows, {len(df.columns)} columns")

    missing = [c for c in args.key_cols if c not in df.columns]
    if missing:
        sys.exit(f"[ERROR] Key columns not found in file: {missing}\n"
                 f"  Available columns: {list(df.columns)}")

    print(f"\nChecking columns: {args.key_cols}")
    print(f"Fuzzy threshold: {args.fuzzy_threshold}")

    annotated = find_duplicates(df, args.key_cols, args.fuzzy_threshold)

    exact_count = (annotated["_dup_type"] == "Exact").sum()
    fuzzy_count = (annotated["_dup_type"] == "Fuzzy").sum()
    total_flagged = exact_count + fuzzy_count

    print(f"\nResults:")
    print(f"  Exact duplicates : {exact_count:,} rows")
    print(f"  Fuzzy matches    : {fuzzy_count:,} rows")
    print(f"  Total flagged    : {total_flagged:,} rows")

    # Build summary
    summary_data = {
        "Metric": ["Total rows", "Exact duplicates", "Fuzzy near-duplicates",
                   "Total flagged", "Clean rows", "Key columns used", "Fuzzy threshold"],
        "Value": [len(df), exact_count, fuzzy_count, total_flagged,
                  len(df) - total_flagged,
                  ", ".join(args.key_cols), args.fuzzy_threshold],
    }

    out = Path(args.output)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        annotated.to_excel(writer, sheet_name="Flagged Data", index=False)
        pd.DataFrame(summary_data).to_excel(writer, sheet_name="Summary", index=False)

    style_output(out, annotated)
    print(f"\nOutput written to: {out.resolve()}")
    print("  Orange rows = exact duplicates")
    print("  Yellow rows = near-duplicates (fuzzy match)")


if __name__ == "__main__":
    main()
  
