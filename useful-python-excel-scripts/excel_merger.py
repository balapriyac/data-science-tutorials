"""
excel_merger.py
Merge multiple Excel / CSV files from a folder into one unified output file.

Dependencies: pandas, openpyxl
Install:      pip install pandas openpyxl

Usage:
    python excel_merger.py --input ./reports --output merged.xlsx
    python excel_merger.py --input ./reports --output merged.xlsx --source-column "Source File"
    python excel_merger.py --input ./reports --output merged.xlsx --sheet "Sheet2" --no-summary
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG (override via CLI flags) ──────────────────────────────────────────
INPUT_FOLDER = "./input"        # Folder containing files to merge
OUTPUT_FILE  = "merged.xlsx"    # Output file path
SHEET_NAME   = None             # Sheet to read from each file; None = first sheet
SOURCE_COLUMN = "Source File"   # Set to None to skip adding source column
ADD_SUMMARY  = True             # Write a summary tab with per-file row counts
# ─────────────────────────────────────────────────────────────────────────────

SUPPORTED = {".xlsx", ".xls", ".csv"}


def read_file(path: Path, sheet: str | None) -> pd.DataFrame:
    """Read an Excel or CSV file into a DataFrame."""
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=str)
    kwargs = {"dtype": str}
    if sheet:
        kwargs["sheet_name"] = sheet
    else:
        kwargs["sheet_name"] = 0
    return pd.read_excel(path, **kwargs)


def merge_files(input_folder: str, output_file: str, sheet: str | None,
                source_col: str | None, add_summary: bool) -> None:
    folder = Path(input_folder)
    if not folder.exists():
        sys.exit(f"[ERROR] Input folder not found: {folder}")

    files = [f for f in sorted(folder.iterdir()) if f.suffix.lower() in SUPPORTED]
    if not files:
        sys.exit(f"[ERROR] No supported files found in {folder}")

    print(f"Found {len(files)} file(s) to merge.\n")

    frames = []
    summary_rows = []
    all_columns = []

    for f in files:
        try:
            df = read_file(f, sheet)
            row_count = len(df)
            if source_col:
                df[source_col] = f.name
            frames.append(df)
            summary_rows.append({"File": f.name, "Rows": row_count, "Columns": df.shape[1]})
            # Track union of all columns (preserving first-seen order)
            for col in df.columns:
                if col not in all_columns:
                    all_columns.append(col)
            print(f"  ✓ {f.name:40s} {row_count:>6} rows")
        except Exception as e:
            print(f"  ✗ {f.name:40s} FAILED — {e}")
            summary_rows.append({"File": f.name, "Rows": "ERROR", "Columns": str(e)})

    if not frames:
        sys.exit("[ERROR] No files could be read.")

    merged = pd.concat(frames, ignore_index=True, sort=False)[all_columns]
    total = len(merged)
    print(f"\nTotal rows after merge: {total:,}")

    # Write output
    out = Path(output_file)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        merged.to_excel(writer, sheet_name="Merged Data", index=False)

        if add_summary:
            summary_df = pd.DataFrame(summary_rows)
            summary_df.loc[len(summary_df)] = {
                "File": f"TOTAL ({len(frames)} files)",
                "Rows": total,
                "Columns": "",
            }
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

    # Style header row
    wb = load_workbook(out)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
        ws.freeze_panes = "A2"
    wb.save(out)

    print(f"\nOutput written to: {out.resolve()}")
    if add_summary:
        print("Summary tab included in output.")


def main():
    parser = argparse.ArgumentParser(description="Merge Excel/CSV files into one.")
    parser.add_argument("--input",         default=INPUT_FOLDER,  help="Input folder path")
    parser.add_argument("--output",        default=OUTPUT_FILE,   help="Output .xlsx file path")
    parser.add_argument("--sheet",         default=SHEET_NAME,    help="Sheet name to read (default: first sheet)")
    parser.add_argument("--source-column", default=SOURCE_COLUMN, help="Column name for source filename (or empty to skip)")
    parser.add_argument("--no-summary",    action="store_true",   help="Skip writing the summary tab")
    args = parser.parse_args()

    merge_files(
        input_folder=args.input,
        output_file=args.output,
        sheet=args.sheet,
        source_col=args.source_column or None,
        add_summary=not args.no_summary,
    )


if __name__ == "__main__":
    main()

