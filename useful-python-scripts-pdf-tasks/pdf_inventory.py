"""
pdf_inventory.py
Scan a folder of PDF files and extract metadata into a single inventory file.

Dependencies: pypdf, pdfplumber, pandas, openpyxl
Install:      pip install pypdf pdfplumber pandas openpyxl

Usage:
    python pdf_inventory.py --input ./documents
    python pdf_inventory.py --input ./documents --output inventory.xlsx --sample-pages 3
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
import pdfplumber
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pypdf import PdfReader
from pypdf.errors import PdfReadError

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FOLDER  = "./documents"
OUTPUT_FILE   = "pdf_inventory.xlsx"
SAMPLE_PAGES  = 3       # Number of pages to sample for text detection
RECURSIVE     = False   # Search subdirectories
# ─────────────────────────────────────────────────────────────────────────────

HEADER_COLOR   = "1F4E79"
ENCRYPT_COLOR  = "FFC7CE"
SCAN_COLOR     = "FFEB9C"
ALT_ROW_COLOR  = "DEEAF1"


def parse_pdf_date(raw: str) -> str:
    """Parse PDF date string (D:YYYYMMDDHHmmSS) to readable format."""
    if not raw:
        return ""
    raw = str(raw).strip().lstrip("D:").replace("'", "")
    for fmt in ("%Y%m%d%H%M%S%z", "%Y%m%d%H%M%S", "%Y%m%d"):
        try:
            return datetime.strptime(raw[:len(fmt.replace("%z",""))], fmt).strftime("%Y-%m-%d %H:%M")
        except ValueError:
            continue
    return raw[:16]


def is_scanned(pdf_path: Path, sample_pages: int) -> bool:
    """Return True if the sampled pages contain no extractable text."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            pages_to_check = pdf.pages[:sample_pages]
            for page in pages_to_check:
                text = page.extract_text() or ""
                if text.strip():
                    return False
        return True
    except Exception:
        return False


def inspect_pdf(pdf_path: Path, sample_pages: int) -> dict:
    result = {
        "filename":      pdf_path.name,
        "path":          str(pdf_path.resolve()),
        "size_kb":       round(pdf_path.stat().st_size / 1024, 2),
        "pages":         None,
        "encrypted":     False,
        "scanned":       False,
        "title":         "",
        "author":        "",
        "creator":       "",
        "producer":      "",
        "created":       "",
        "modified":      "",
        "pdf_version":   "",
        "error":         "",
    }

    try:
        reader = PdfReader(pdf_path)

        if reader.is_encrypted:
            result["encrypted"] = True
            # Try empty password
            try:
                reader.decrypt("")
            except Exception:
                result["error"] = "Encrypted — could not open"
                return result

        result["pages"]       = len(reader.pages)
        result["pdf_version"] = reader.pdf_header if hasattr(reader, "pdf_header") else ""

        meta = reader.metadata or {}
        result["title"]    = str(meta.get("/Title",    "")).strip()
        result["author"]   = str(meta.get("/Author",   "")).strip()
        result["creator"]  = str(meta.get("/Creator",  "")).strip()
        result["producer"] = str(meta.get("/Producer", "")).strip()
        result["created"]  = parse_pdf_date(str(meta.get("/CreationDate", "")))
        result["modified"] = parse_pdf_date(str(meta.get("/ModDate",      "")))

        result["scanned"] = is_scanned(pdf_path, sample_pages)

    except PdfReadError as e:
        result["error"] = f"PdfReadError: {e}"
    except Exception as e:
        result["error"] = str(e)

    return result


def style_wb(path: Path, records: list[dict]) -> None:
    wb = load_workbook(path)
    ws = wb.active

    header_fill  = PatternFill("solid", fgColor=HEADER_COLOR)
    encrypt_fill = PatternFill("solid", fgColor=ENCRYPT_COLOR)
    scan_fill    = PatternFill("solid", fgColor=SCAN_COLOR)
    alt_fill     = PatternFill("solid", fgColor=ALT_ROW_COLOR)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    headers = [cell.value for cell in ws[1]]
    enc_col = headers.index("encrypted") + 1 if "encrypted" in headers else None
    scn_col = headers.index("scanned")   + 1 if "scanned"   in headers else None

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        is_enc = ws.cell(row=row_idx + 2, column=enc_col).value if enc_col else False
        is_scn = ws.cell(row=row_idx + 2, column=scn_col).value if scn_col else False

        if is_enc:
            fill = encrypt_fill
        elif is_scn:
            fill = scan_fill
        elif row_idx % 2 == 0:
            fill = alt_fill
        else:
            fill = None

        if fill:
            for cell in row:
                cell.fill = fill

    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22
    ws.freeze_panes = "A2"

    # Summary sheet
    if "Summary" in wb.sheetnames:
        ws_sum = wb["Summary"]
        for cell in ws_sum[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Generate a metadata inventory of PDF files.")
    parser.add_argument("--input",        default=INPUT_FOLDER,
                        help="Folder containing PDF files")
    parser.add_argument("--output",       default=OUTPUT_FILE)
    parser.add_argument("--sample-pages", type=int, default=SAMPLE_PAGES,
                        help="Pages to sample for scanned-image detection")
    parser.add_argument("--recursive",    action="store_true", default=RECURSIVE,
                        help="Search subdirectories")
    args = parser.parse_args()

    folder = Path(args.input)
    if not folder.exists():
        sys.exit(f"[ERROR] Folder not found: {folder}")

    glob_pattern = "**/*.pdf" if args.recursive else "*.pdf"
    pdfs = sorted(folder.glob(glob_pattern))
    if not pdfs:
        sys.exit(f"[ERROR] No PDF files found in: {folder}")

    print(f"Found {len(pdfs):,} PDF file(s)\n")

    records = []
    for i, pdf_path in enumerate(pdfs, 1):
        print(f"  [{i}/{len(pdfs)}] {pdf_path.name}")
        record = inspect_pdf(pdf_path, args.sample_pages)
        records.append(record)
        flags = []
        if record["encrypted"]: flags.append("ENCRYPTED")
        if record["scanned"]:   flags.append("SCANNED")
        if record["error"]:     flags.append(f"ERROR: {record['error']}")
        if flags:
            print(f"    ⚠ {', '.join(flags)}")

    df = pd.DataFrame(records)

    # Summary stats
    total_size = df["size_kb"].sum()
    summary = pd.DataFrame([
        {"Metric": "Total files",       "Value": len(df)},
        {"Metric": "Total size (KB)",   "Value": round(total_size, 2)},
        {"Metric": "Total size (MB)",   "Value": round(total_size / 1024, 2)},
        {"Metric": "Total pages",       "Value": df["pages"].sum()},
        {"Metric": "Avg pages/file",    "Value": round(df["pages"].mean(), 1)},
        {"Metric": "Encrypted files",   "Value": df["encrypted"].sum()},
        {"Metric": "Scanned (image) files", "Value": df["scanned"].sum()},
        {"Metric": "Files with errors", "Value": (df["error"] != "").sum()},
    ])

    print(f"\nTotal files : {len(df)}")
    print(f"Total pages : {df['pages'].sum()}")
    print(f"Encrypted   : {df['encrypted'].sum()}")
    print(f"Scanned     : {df['scanned'].sum()}")

    out = Path(args.output)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Inventory", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)

    style_wb(out, records)
    print(f"\nOutput written to: {out.resolve()}")
    print("  Yellow rows = scanned/image PDFs (no extractable text)")
    print("  Red rows    = encrypted files")


if __name__ == "__main__":
    main()
  
