"""
Transforms analyzed data into polished Excel reports with professional
formatting, conditional styling, and summary statistics.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from datetime import datetime


def format_report(df, output_file, title="Analysis Report"):
    """
    Creates a professionally formatted Excel report from a DataFrame.
    
    Args:
        df: pandas DataFrame with analysis results
        output_file: path to save the formatted Excel file
        title: report title to display at the top
    """
    # Save initial DataFrame to Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Report', index=False, startrow=2)
    
    # Load workbook for formatting
    wb = load_workbook(output_file)
    ws = wb['Report']
    
    # Add title
    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78")
    
    # Add timestamp
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A2'].font = Font(size=10, italic=True)
    
    # Format headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col in range(1, len(df.columns) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Format data rows
    for row in range(3, len(df) + 4):
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Format numbers
            col_name = df.columns[col-1].lower()
            if cell.row > 3:  # Skip header
                if any(keyword in col_name for keyword in ['price', 'cost', 'revenue', 'amount']):
                    cell.number_format = '$#,##0.00'
                elif any(keyword in col_name for keyword in ['percent', 'rate', '%']):
                    cell.number_format = '0.00%'
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    # Add conditional formatting to numeric columns
    for col_idx, col_name in enumerate(df.columns, start=1):
        if df[col_name].dtype in ['int64', 'float64']:
            col_letter = get_column_letter(col_idx)
            data_range = f"{col_letter}4:{col_letter}{len(df) + 3}"
            
            # Color scale: red (low) to green (high)
            color_scale = ColorScaleRule(
                start_type='min', start_color='F8696B',
                end_type='max', end_color='63BE7B'
            )
            ws.conditional_formatting.add(data_range, color_scale)
    
    # Auto-adjust column widths
    for col in range(1, len(df.columns) + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in range(3, len(df) + 4):
            cell_value = str(ws.cell(row=row, column=col).value)
            max_length = max(max_length, len(cell_value))
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add summary row if numeric columns exist
    numeric_cols = df.select_dtypes(include=['int64', 'float64']).columns
    if len(numeric_cols) > 0:
        summary_row = len(df) + 4
        ws.cell(row=summary_row, column=1).value = "TOTALS"
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        for col_idx, col_name in enumerate(df.columns, start=1):
            if col_name in numeric_cols:
                col_letter = get_column_letter(col_idx)
                ws.cell(row=summary_row, column=col_idx).value = f"=SUM({col_letter}4:{col_letter}{len(df) + 3})"
                ws.cell(row=summary_row, column=col_idx).font = Font(bold=True)
                ws.cell(row=summary_row, column=col_idx).fill = PatternFill(
                    start_color="E7E6E6", end_color="E7E6E6", fill_type="solid"
                )
    
    wb.save(output_file)
    print(f"✓ Formatted report saved to {output_file}")


# Example usage
if __name__ == "__main__":
    # Sample data
    data = {
        'Region': ['North', 'South', 'East', 'West', 'Central'],
        'Revenue': [125000, 98000, 145000, 112000, 89000],
        'Growth_Rate': [0.15, 0.08, 0.22, 0.11, 0.05],
        'Units_Sold': [1250, 980, 1450, 1120, 890]
    }
    df = pd.DataFrame(data)
    
    format_report(df, "formatted_report.xlsx", "Q3 Regional Performance")

