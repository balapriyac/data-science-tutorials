# =====================================================
# SCRIPT 2: Database Query Result Formatter & Validator
# =====================================================

import sqlite3
import pandas as pd
from sqlalchemy import create_engine
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill

class DatabaseReportGenerator:
    def __init__(self, connection_string):
        self.engine = create_engine(connection_string)
    
    def execute_and_format(self, query, output_file=None):
        """Execute query and format results for presentation"""
        try:
            # Execute query
            df = pd.read_sql(query, self.engine)
            
            # Validate data
            validation_report = self._validate_data(df)
            
            # Format for presentation
            formatted_df = self._format_dataframe(df)
            
            # Export to Excel if requested
            if output_file:
                self._export_to_excel(formatted_df, output_file, validation_report)
            
            return {
                'data': formatted_df,
                'validation': validation_report,
                'summary': self._generate_summary(df)
            }
            
        except Exception as e:
            return {'error': str(e)}
    
    def _validate_data(self, df):
        """Validate data for common issues"""
        validation = {
            'row_count': len(df),
            'column_count': len(df.columns),
            'missing_values': df.isnull().sum().to_dict(),
            'duplicates': df.duplicated().sum(),
            'data_types': df.dtypes.to_dict()
        }
        
        # Flag potential issues
        issues = []
        if validation['duplicates'] > 0:
            issues.append(f"Found {validation['duplicates']} duplicate rows")
        
        missing_pct = df.isnull().sum() / len(df) * 100
        high_missing = missing_pct[missing_pct > 20]
        if not high_missing.empty:
            issues.append(f"High missing values in: {', '.join(high_missing.index)}")
        
        validation['issues'] = issues
        return validation
    
    def _format_dataframe(self, df):
        """Format dataframe for presentation"""
        formatted = df.copy()
        
        # Format numeric columns
        for col in formatted.select_dtypes(include=['float64', 'int64']).columns:
            if formatted[col].max() > 1000:
                formatted[col] = formatted[col].apply(lambda x: f"{x:,.0f}")
            else:
                formatted[col] = formatted[col].apply(lambda x: f"{x:.2f}")
        
        # Format date columns
        for col in formatted.select_dtypes(include=['datetime64']).columns:
            formatted[col] = formatted[col].dt.strftime('%Y-%m-%d')
        
        return formatted
    
    def _export_to_excel(self, df, filename, validation):
        """Export to Excel with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Query Results"
        
        # Add headers with formatting
        headers = list(df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Add data
        for row in dataframe_to_rows(df, index=False, header=False):
            ws.append(row)
        
        # Add validation sheet
        validation_ws = wb.create_sheet("Data Validation")
        validation_ws.append(["Metric", "Value"])
        for key, value in validation.items():
            if key != 'issues':
                validation_ws.append([key, str(value)])
        
        wb.save(filename)
    
    def _generate_summary(self, df):
        """Generate executive summary"""
        numeric_cols = df.select_dtypes(include=[np.number]).columns
        summary = {}
        
        for col in numeric_cols:
            summary[col] = {
                'mean': df[col].mean(),
                'median': df[col].median(),
                'std': df[col].std(),
                'min': df[col].min(),
                'max': df[col].max()
            }
        
        return summary

# Example usage
db_reporter = DatabaseReportGenerator('sqlite:///sample.db')
query = "SELECT * FROM sales_data WHERE date >= '2024-01-01'"
results = db_reporter.execute_and_format(query, 'sales_report.xlsx')

