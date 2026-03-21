"""
ts_compare.py
Compare multiple time series: correlation, lag analysis, and summary statistics.

Dependencies: pandas, openpyxl, matplotlib, scipy
Install:      pip install pandas openpyxl matplotlib scipy

Usage:
    python ts_compare.py --input data.csv --datetime-col "Date" --value-cols "Sales" "Traffic" "Conversions"
    python ts_compare.py --input data.csv --datetime-col "Date" --value-cols "A" "B" "C" --freq MS --max-lag 6
    python ts_compare.py --input data.csv --datetime-col "Date" --value-cols "Revenue" "Cost" --plot
"""

import argparse
import sys
from itertools import combinations
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from scipy.stats import pearsonr, spearmanr

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FILE   = "data.csv"
OUTPUT_FILE  = "comparison_report.xlsx"
DATETIME_COL = "Date"
VALUE_COLS   = []        # Empty = all numeric columns
FREQ         = "MS"      # Resample frequency for alignment
MAX_LAG      = 12        # Maximum lag periods for cross-correlation
TOP_PAIRS    = 5         # Number of top correlated pairs to chart
SAVE_PLOTS   = True
# ─────────────────────────────────────────────────────────────────────────────

HEADER_COLOR = "1F4E79"
PALETTE = ["#1F4E79", "#2E86AB", "#70AD47", "#FF6B35", "#FFC300",
           "#A23B72", "#C73E1D", "#3B1F2B", "#44BBA4", "#E94F37"]


def load_and_align(path: Path, datetime_col: str, value_cols: list[str],
                   freq: str) -> pd.DataFrame:
    df = pd.read_csv(path) if path.suffix.lower() == ".csv" else pd.read_excel(path)

    if datetime_col not in df.columns:
        sys.exit(f"[ERROR] Datetime column '{datetime_col}' not found.")

    df[datetime_col] = pd.to_datetime(df[datetime_col], infer_datetime_format=True, errors="coerce")
    df.dropna(subset=[datetime_col], inplace=True)
    df.set_index(datetime_col, inplace=True)
    df.sort_index(inplace=True)

    numeric_cols = list(df.select_dtypes(include="number").columns)
    cols = value_cols if value_cols else numeric_cols
    missing = [c for c in cols if c not in df.columns]
    if missing:
        sys.exit(f"[ERROR] Columns not found: {missing}")

    aligned = df[cols].resample(freq).mean()
    aligned = aligned.interpolate(method="time")
    return aligned


def compute_correlations(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return Pearson and Spearman correlation matrices."""
    pairs = list(combinations(df.columns, 2))
    rows = []
    for a, b in pairs:
        s_a = df[a].dropna()
        s_b = df[b].dropna()
        idx = s_a.index.intersection(s_b.index)
        if len(idx) < 3:
            continue
        pearson_r,  pearson_p  = pearsonr(s_a[idx],  s_b[idx])
        spearman_r, spearman_p = spearmanr(s_a[idx], s_b[idx])
        rows.append({
            "Series A":   a,
            "Series B":   b,
            "Pearson r":  round(pearson_r,  4),
            "Pearson p":  round(pearson_p,  4),
            "Spearman r": round(spearman_r, 4),
            "Spearman p": round(spearman_p, 4),
            "N":          len(idx),
        })
    corr_df = pd.DataFrame(rows).sort_values("Pearson r", key=abs, ascending=False)

    # Wide correlation matrix (Pearson)
    matrix = df.corr(method="pearson").round(4)
    return corr_df, matrix


def compute_lag_analysis(df: pd.DataFrame, max_lag: int) -> pd.DataFrame:
    """Cross-correlation: find the lag at which each pair peaks."""
    pairs = list(combinations(df.columns, 2))
    rows = []
    for a, b in pairs:
        s_a = (df[a] - df[a].mean()) / df[a].std()
        s_b = (df[b] - df[b].mean()) / df[b].std()
        aligned = pd.concat([s_a, s_b], axis=1).dropna()
        if len(aligned) < max_lag * 2:
            continue
        xcorr = np.correlate(aligned.iloc[:, 0], aligned.iloc[:, 1], mode="full")
        xcorr /= len(aligned)
        center = len(xcorr) // 2
        lag_range = range(-max_lag, max_lag + 1)
        xcorr_window = xcorr[center - max_lag: center + max_lag + 1]
        peak_idx = np.argmax(np.abs(xcorr_window))
        peak_lag  = list(lag_range)[peak_idx]
        peak_corr = xcorr_window[peak_idx]
        rows.append({
            "Series A":     a,
            "Series B":     b,
            "Peak Lag":     peak_lag,
            "Peak XCorr":   round(float(peak_corr), 4),
            "Interpretation": (
                f"{a} leads {b} by {abs(peak_lag)} period(s)" if peak_lag < 0
                else f"{b} leads {a} by {abs(peak_lag)} period(s)" if peak_lag > 0
                else "Contemporaneous (no lag)"
            ),
        })
    return pd.DataFrame(rows).sort_values("Peak XCorr", key=abs, ascending=False)


def compute_summary_stats(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for col in df.columns:
        s = df[col].dropna()
        slope = np.polyfit(np.arange(len(s)), s.values, 1)[0]
        rows.append({
            "Series":          col,
            "Count":           len(s),
            "Mean":            round(s.mean(),  4),
            "Std Dev":         round(s.std(),   4),
            "Min":             round(s.min(),   4),
            "Max":             round(s.max(),   4),
            "Median":          round(s.median(),4),
            "Trend Slope":     round(slope,     6),
            "Trend Direction": "Upward" if slope > 0 else "Downward",
        })
    return pd.DataFrame(rows)


def save_pair_charts(df: pd.DataFrame, corr_df: pd.DataFrame,
                     top_n: int, out_dir: Path) -> None:
    top_pairs = corr_df.head(top_n)[["Series A", "Series B"]].values.tolist()
    for i, (a, b) in enumerate(top_pairs):
        fig, ax1 = plt.subplots(figsize=(13, 4))
        color_a, color_b = PALETTE[0], PALETTE[1]

        ax1.plot(df.index, df[a], color=color_a, linewidth=1.0, label=a)
        ax1.set_ylabel(a, color=color_a, fontsize=9)
        ax1.tick_params(axis="y", labelcolor=color_a)

        ax2 = ax1.twinx()
        ax2.plot(df.index, df[b], color=color_b, linewidth=1.0, linestyle="--", label=b)
        ax2.set_ylabel(b, color=color_b, fontsize=9)
        ax2.tick_params(axis="y", labelcolor=color_b)

        r_val = corr_df.loc[(corr_df["Series A"] == a) & (corr_df["Series B"] == b), "Pearson r"]
        r_str = f"r = {r_val.values[0]:.3f}" if not r_val.empty else ""
        ax1.set_title(f"{a} vs {b}  {r_str}", fontsize=11)
        ax1.spines[["top"]].set_visible(False)
        ax2.spines[["top"]].set_visible(False)

        lines1, labels1 = ax1.get_legend_handles_labels()
        lines2, labels2 = ax2.get_legend_handles_labels()
        ax1.legend(lines1 + lines2, labels1 + labels2, fontsize=8, loc="upper left")

        plt.tight_layout()
        chart_path = out_dir / f"pair_{a.replace(' ','_')}_vs_{b.replace(' ','_')}.png"
        plt.savefig(chart_path, dpi=150, bbox_inches="tight")
        plt.close()
        print(f"  Chart saved: {chart_path.name}")


def style_wb(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.freeze_panes = "A2"
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Compare multiple time series columns.")
    parser.add_argument("--input",        default=INPUT_FILE)
    parser.add_argument("--output",       default=OUTPUT_FILE)
    parser.add_argument("--datetime-col", default=DATETIME_COL)
    parser.add_argument("--value-cols",   nargs="*", default=VALUE_COLS,
                        help="Columns to compare (default: all numeric)")
    parser.add_argument("--freq",         default=FREQ, help="Alignment frequency (MS, D, W, ...)")
    parser.add_argument("--max-lag",      type=int, default=MAX_LAG, help="Max lag for cross-correlation")
    parser.add_argument("--top-pairs",    type=int, default=TOP_PAIRS, help="Top correlated pairs to chart")
    parser.add_argument("--plot",         action="store_true", default=SAVE_PLOTS)
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    print(f"Reading and aligning: {src}")
    df = load_and_align(src, args.datetime_col, args.value_cols, args.freq)
    print(f"  {len(df.columns)} series | {len(df):,} aligned periods | freq: {args.freq}")
    print(f"  Columns: {list(df.columns)}")

    print("\nComputing correlations...")
    corr_df, corr_matrix = compute_correlations(df)

    print("Computing lag analysis...")
    lag_df = compute_lag_analysis(df, args.max_lag)

    print("Computing summary statistics...")
    stats_df = compute_summary_stats(df)

    out = Path(args.output)
    aligned_out = df.reset_index()
    aligned_out.rename(columns={"index": args.datetime_col}, errors="ignore", inplace=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        aligned_out.to_excel(writer,     sheet_name="Aligned Series",   index=False)
        stats_df.to_excel(writer,        sheet_name="Summary Stats",    index=False)
        corr_df.to_excel(writer,         sheet_name="Pairwise Corr",    index=False)
        corr_matrix.to_excel(writer,     sheet_name="Corr Matrix",      index=True)
        lag_df.to_excel(writer,          sheet_name="Lag Analysis",     index=False)

    style_wb(out)

    if args.plot and not corr_df.empty:
        save_pair_charts(df, corr_df, args.top_pairs, Path(args.output).parent)

    print(f"\nOutput written to: {out.resolve()}")
    print("  Tabs: Aligned Series | Summary Stats | Pairwise Corr | Corr Matrix | Lag Analysis")
    if not corr_df.empty:
        print(f"\nTop correlated pairs:")
        for _, row in corr_df.head(3).iterrows():
            print(f"  {row['Series A']} × {row['Series B']}: r={row['Pearson r']}")


if __name__ == "__main__":
    main()
  
