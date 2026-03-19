"""
ts_forecast.py
Fit a SARIMA model and generate a forward forecast with confidence intervals.

Dependencies: pandas, openpyxl, statsmodels, matplotlib
Install:      pip install pandas openpyxl statsmodels matplotlib

Usage:
    python ts_forecast.py --input data.csv --datetime-col "Date" --value-col "Sales" --periods 12
    python ts_forecast.py --input data.csv --datetime-col "Date" --value-col "Revenue" --periods 6 --auto-order
    python ts_forecast.py --input data.csv --datetime-col "Date" --value-col "Sales" --order 1 1 1 --seasonal-order 1 1 1 12

SARIMA notation: (p, d, q) x (P, D, Q, s)
    p/P = AR order        d/D = differencing    q/Q = MA order    s = seasonal period
"""

import argparse
import itertools
import sys
import warnings
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from statsmodels.tsa.statespace.sarimax import SARIMAX

warnings.filterwarnings("ignore")

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FILE      = "data.csv"
OUTPUT_FILE     = "forecast.xlsx"
DATETIME_COL    = "Date"
VALUE_COL       = "Value"
FREQ            = "MS"          # Series frequency for SARIMA (MS, D, W, M, Q, A)
FORECAST_PERIODS = 12           # Periods to forecast ahead
TEST_PERIODS    = 6             # Hold-out periods for accuracy evaluation
AUTO_ORDER      = False         # Grid-search for best ARIMA orders
# Manual orders (used when AUTO_ORDER = False)
ORDER           = (1, 1, 1)     # (p, d, q)
SEASONAL_ORDER  = (1, 1, 1, 12) # (P, D, Q, s)  — set s=0 to disable seasonal component
SAVE_PLOT       = True
# ─────────────────────────────────────────────────────────────────────────────

HEADER_COLOR = "1F4E79"
FORECAST_COLOR = "70AD47"


def load_series(path: Path, datetime_col: str, value_col: str, freq: str) -> pd.Series:
    df = pd.read_csv(path) if path.suffix.lower() == ".csv" else pd.read_excel(path)
    for col in [datetime_col, value_col]:
        if col not in df.columns:
            sys.exit(f"[ERROR] Column '{col}' not found. Available: {list(df.columns)}")
    df[datetime_col] = pd.to_datetime(df[datetime_col], infer_datetime_format=True, errors="coerce")
    df.dropna(subset=[datetime_col], inplace=True)
    df.set_index(datetime_col, inplace=True)
    df.sort_index(inplace=True)
    series = pd.to_numeric(df[value_col], errors="coerce").dropna()
    series = series.resample(freq).mean().interpolate(method="time")
    return series


def auto_select_order(series: pd.Series, seasonal_period: int) -> tuple[tuple, tuple]:
    """Light grid search over ARIMA and seasonal orders, return best by AIC."""
    print("  Running auto order selection (this may take a moment)...")
    best_aic   = np.inf
    best_order = (1, 1, 1)
    best_sorder = (0, 0, 0, seasonal_period)

    p_range = [0, 1, 2]
    d_range = [0, 1]
    q_range = [0, 1, 2]
    P_range = [0, 1]
    D_range = [0, 1]
    Q_range = [0, 1]

    for p, d, q in itertools.product(p_range, d_range, q_range):
        for P, D, Q in itertools.product(P_range, D_range, Q_range):
            try:
                model = SARIMAX(
                    series,
                    order=(p, d, q),
                    seasonal_order=(P, D, Q, seasonal_period),
                    enforce_stationarity=False,
                    enforce_invertibility=False,
                )
                fit = model.fit(disp=False)
                if fit.aic < best_aic:
                    best_aic    = fit.aic
                    best_order  = (p, d, q)
                    best_sorder = (P, D, Q, seasonal_period)
            except Exception:
                continue

    print(f"  Best order: {best_order} x {best_sorder} (AIC={best_aic:.2f})")
    return best_order, best_sorder


def fit_and_forecast(series: pd.Series, order: tuple, seasonal_order: tuple,
                     forecast_periods: int, test_periods: int):
    """Fit on train split, evaluate on test, refit on full series, forecast."""
    train = series.iloc[:-test_periods]
    test  = series.iloc[-test_periods:]

    # Fit on training data
    train_model = SARIMAX(
        train, order=order, seasonal_order=seasonal_order,
        enforce_stationarity=False, enforce_invertibility=False,
    )
    train_fit = train_model.fit(disp=False)

    # Predict on test period
    pred_obj = train_fit.get_prediction(
        start=test.index[0], end=test.index[-1], dynamic=False
    )
    test_pred   = pred_obj.predicted_mean
    mae  = np.mean(np.abs(test.values - test_pred.values))
    rmse = np.sqrt(np.mean((test.values - test_pred.values) ** 2))
    mape = np.mean(np.abs((test.values - test_pred.values) / test.values)) * 100

    # Refit on full series
    full_model = SARIMAX(
        series, order=order, seasonal_order=seasonal_order,
        enforce_stationarity=False, enforce_invertibility=False,
    )
    full_fit = full_model.fit(disp=False)

    # Forward forecast
    forecast_obj = full_fit.get_forecast(steps=forecast_periods)
    forecast_mean = forecast_obj.predicted_mean
    conf_int = forecast_obj.conf_int(alpha=0.05)

    return {
        "train":          train,
        "test":           test,
        "test_pred":      test_pred,
        "forecast_mean":  forecast_mean,
        "conf_int":       conf_int,
        "mae":            mae,
        "rmse":           rmse,
        "mape":           mape,
        "aic":            full_fit.aic,
        "bic":            full_fit.bic,
    }


def save_forecast_plot(results: dict, value_col: str, order: tuple,
                       seasonal_order: tuple, out_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(14, 5))

    ax.plot(results["train"].index, results["train"].values,
            color="#1F4E79", linewidth=1.0, label="Historical (train)")
    ax.plot(results["test"].index, results["test"].values,
            color="#1F4E79", linewidth=1.0, linestyle="--", label="Historical (test)")
    ax.plot(results["test_pred"].index, results["test_pred"].values,
            color="#FF6B35", linewidth=1.0, label="Test prediction")
    ax.plot(results["forecast_mean"].index, results["forecast_mean"].values,
            color="#70AD47", linewidth=1.5, label="Forecast")
    ax.fill_between(
        results["conf_int"].index,
        results["conf_int"].iloc[:, 0],
        results["conf_int"].iloc[:, 1],
        color="#70AD47", alpha=0.15, label="95% confidence interval",
    )

    ax.axvline(results["test"].index[0], color="#aaa", linestyle=":", linewidth=1)
    ax.axvline(results["forecast_mean"].index[0], color="#ccc", linestyle=":", linewidth=1)
    ax.set_title(
        f"SARIMA Forecast — {value_col}  "
        f"(order={order}, seasonal={seasonal_order})",
        fontsize=11,
    )
    ax.set_xlabel("Date")
    ax.set_ylabel(value_col)
    ax.legend(fontsize=8)
    ax.spines[["top", "right"]].set_visible(False)
    plt.tight_layout()
    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Chart saved: {out_path.name}")


def style_wb(path: Path) -> None:
    wb = load_workbook(path)
    fc_fill     = PatternFill("solid", fgColor=FORECAST_COLOR)
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        if sname == "Forecast":
            for row in ws.iter_rows(min_row=2):
                if row[1].value == "Forecast":
                    for cell in row:
                        cell.fill = fc_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.freeze_panes = "A2"
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="SARIMA forecast for a time series column.")
    parser.add_argument("--input",           default=INPUT_FILE)
    parser.add_argument("--output",          default=OUTPUT_FILE)
    parser.add_argument("--datetime-col",    default=DATETIME_COL)
    parser.add_argument("--value-col",       default=VALUE_COL)
    parser.add_argument("--freq",            default=FREQ, help="Pandas frequency string")
    parser.add_argument("--periods",         type=int, default=FORECAST_PERIODS, help="Periods to forecast")
    parser.add_argument("--test-periods",    type=int, default=TEST_PERIODS, help="Hold-out periods for validation")
    parser.add_argument("--auto-order",      action="store_true", default=AUTO_ORDER)
    parser.add_argument("--order",           type=int, nargs=3, default=list(ORDER), metavar=("p","d","q"))
    parser.add_argument("--seasonal-order",  type=int, nargs=4, default=list(SEASONAL_ORDER), metavar=("P","D","Q","s"))
    parser.add_argument("--plot",            action="store_true", default=SAVE_PLOT)
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    print(f"Reading: {src}")
    series = load_series(src, args.datetime_col, args.value_col, args.freq)
    print(f"  {len(series):,} data points | freq: {args.freq}")
    print(f"  Date range: {series.index.min()} → {series.index.max()}")

    order          = tuple(args.order)
    seasonal_order = tuple(args.seasonal_order)

    if args.auto_order:
        s = seasonal_order[3] if seasonal_order[3] > 0 else 12
        order, seasonal_order = auto_select_order(series, s)
    else:
        print(f"\nUsing order={order}, seasonal_order={seasonal_order}")

    print(f"\nFitting model | test_periods={args.test_periods} | forecast={args.periods}")
    results = fit_and_forecast(series, order, seasonal_order, args.periods, args.test_periods)

    print(f"\nValidation accuracy (test period of {args.test_periods} periods):")
    print(f"  MAE  = {results['mae']:.4f}")
    print(f"  RMSE = {results['rmse']:.4f}")
    print(f"  MAPE = {results['mape']:.2f}%")
    print(f"  AIC  = {results['aic']:.2f}  |  BIC = {results['bic']:.2f}")

    # Build forecast DataFrame
    fc_df = pd.DataFrame({
        "Date":              results["forecast_mean"].index,
        "Type":              "Forecast",
        "Value":             results["forecast_mean"].values.round(4),
        "Lower 95%":         results["conf_int"].iloc[:, 0].values.round(4),
        "Upper 95%":         results["conf_int"].iloc[:, 1].values.round(4),
    })

    # Historical + test predictions
    hist_df = pd.DataFrame({
        "Date":              series.index,
        "Type":              ["Historical"] * (len(series) - args.test_periods) +
                             ["Test (actual)"] * args.test_periods,
        "Value":             series.values.round(4),
        "Lower 95%":         np.nan,
        "Upper 95%":         np.nan,
    })

    test_pred_df = pd.DataFrame({
        "Date":      results["test_pred"].index,
        "Type":      "Test (predicted)",
        "Value":     results["test_pred"].values.round(4),
        "Lower 95%": np.nan,
        "Upper 95%": np.nan,
    })

    combined = pd.concat([hist_df, test_pred_df, fc_df], ignore_index=True)

    metrics_df = pd.DataFrame([
        {"Metric": "Order (p,d,q)",         "Value": str(order)},
        {"Metric": "Seasonal order (P,D,Q,s)","Value": str(seasonal_order)},
        {"Metric": "Forecast periods",       "Value": args.periods},
        {"Metric": "Test periods",           "Value": args.test_periods},
        {"Metric": "MAE",                    "Value": round(results["mae"],  4)},
        {"Metric": "RMSE",                   "Value": round(results["rmse"], 4)},
        {"Metric": "MAPE (%)",               "Value": round(results["mape"], 2)},
        {"Metric": "AIC",                    "Value": round(results["aic"],  2)},
        {"Metric": "BIC",                    "Value": round(results["bic"],  2)},
    ])

    out = Path(args.output)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        combined.to_excel(writer,    sheet_name="Forecast",        index=False)
        fc_df.to_excel(writer,       sheet_name="Forecast Only",   index=False)
        metrics_df.to_excel(writer,  sheet_name="Model Metrics",   index=False)

    style_wb(out)

    if args.plot:
        plot_path = Path(args.output).with_suffix(".png")
        save_forecast_plot(results, args.value_col, order, seasonal_order, plot_path)

    print(f"\nOutput written to: {out.resolve()}")
    print("  Green rows = forecast values")


if __name__ == "__main__":
    main()

