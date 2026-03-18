"""
ts_anomaly_detector.py
Detect anomalies in time series data using z-score, IQR, or rolling statistics.

Dependencies: pandas, openpyxl, matplotlib
Install:      pip install pandas openpyxl matplotlib

Usage:
    python ts_anomaly_detector.py --input data.csv --datetime-col "Date" --value-cols "Sales"
    python ts_anomaly_detector.py --input data.csv --datetime-col "ts" --value-cols "Temp" "Pressure" --method rolling
    python ts_anomaly_detector.py --input data.csv --datetime-col "Date" --value-cols "Revenue" --method all --plot
"""

import argparse
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FILE    = "data.csv"
OUTPUT_FILE   = "anomalies_flagged.xlsx"
DATETIME_COL  = "Date"
VALUE_COLS    = []           # Empty = all numeric columns
METHOD        = "zscore"     # zscore | iqr | rolling | all
ZSCORE_THRESH = 3.0          # Standard deviations
IQR_FACTOR    = 1.5          # IQR multiplier
ROLLING_WIN   = 14           # Rolling window (periods)
ROLLING_STD   = 3.0          # Rolling std deviation threshold
SAVE_PLOTS    = False
# ─────────────────────────────────────────────────────────────────────────────

ANOMALY_COLOR = "FF4500"
HEADER_COLOR  = "1F4E79"


def detect_zscore(series: pd.Series, threshold: float) -> pd.Series:
    mean, std = series.mean(), series.std()
    if std == 0:
        return pd.Series(False, index=series.index)
    return ((series - mean) / std).abs() > threshold


def detect_iqr(series: pd.Series, factor: float) -> pd.Series:
    q1, q3 = series.quantile(0.25), series.quantile(0.75)
    iqr = q3 - q1
    return (series < q1 - factor * iqr) | (series > q3 + factor * iqr)


def detect_rolling(series: pd.Series, window: int, n_std: float) -> pd.Series:
    rolling_mean = series.rolling(window, center=True, min_periods=1).mean()
    rolling_std  = series.rolling(window, center=True, min_periods=1).std()
    rolling_std.replace(0, np.nan, inplace=True)
    z = (series - rolling_mean) / rolling_std
    return z.abs() > n_std


def run_detection(series: pd.Series, method: str, cfg: dict) -> pd.Series:
    """Return boolean Series marking anomalous points."""
    flags = pd.Series(False, index=series.index)
    if method in ("zscore", "all"):
        flags |= detect_zscore(series, cfg["zscore_thresh"])
    if method in ("iqr", "all"):
        flags |= detect_iqr(series, cfg["iqr_factor"])
    if method in ("rolling", "all"):
        flags |= detect_rolling(series, cfg["rolling_win"], cfg["rolling_std"])
    return flags


def save_plot(series: pd.Series, flags: pd.Series, col: str, out_dir: Path) -> None:
    fig, ax = plt.subplots(figsize=(14, 4))
    ax.plot(series.index, series.values, color="#1F4E79", linewidth=0.9, label=col)
    anomaly_vals = series[flags]
    ax.scatter(anomaly_vals.index, anomaly_vals.values,
               color="#FF4500", s=40, zorder=5, label="Anomaly")
    ax.set_title(f"Anomaly Detection — {col}", fontsize=12)
    ax.set_xlabel("Date")
    ax.set_ylabel(col)
    ax.legend()
    ax.xaxis.set_major_formatter(mdates.AutoDateFormatter(mdates.AutoDateLocator()))
    fig.autofmt_xdate()
    plt.tight_layout()
    plot_path = out_dir / f"anomalies_{col.replace(' ', '_')}.png"
    plt.savefig(plot_path, dpi=150)
    plt.close()
    print(f"  Chart saved: {plot_path.name}")


def style_wb(path: Path, flag_col_indices: dict) -> None:
    """flag_col_indices: sheet_name → list of column indices (1-based) with anomaly flags."""
    wb = load_workbook(path)
    anomaly_fill = PatternFill("solid", fgColor=ANOMALY_COLOR)
    header_fill  = PatternFill("solid", fgColor=HEADER_COLOR)

    for sname in wb.sheetnames:
        ws = wb[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Highlight anomaly rows in main sheet
        if sname == "Flagged Data":
            flag_cols = flag_col_indices.get(sname, [])
            for row in ws.iter_rows(min_row=2):
                is_anomaly = any(
                    ws.cell(row=row[0].row, column=c).value == "YES"
                    for c in flag_cols
                )
                if is_anomaly:
                    for cell in row:
                        cell.fill = anomaly_fill

        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        ws.freeze_panes = "A2"

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Detect anomalies in time series columns.")
    parser.add_argument("--input",          default=INPUT_FILE)
    parser.add_argument("--output",         default=OUTPUT_FILE)
    parser.add_argument("--datetime-col",   default=DATETIME_COL)
    parser.add_argument("--value-cols",     nargs="*", default=VALUE_COLS,
                        help="Columns to analyse (default: all numeric)")
    parser.add_argument("--method",         default=METHOD,
                        choices=["zscore", "iqr", "rolling", "all"])
    parser.add_argument("--zscore-thresh",  type=float, default=ZSCORE_THRESH)
    parser.add_argument("--iqr-factor",     type=float, default=IQR_FACTOR)
    parser.add_argument("--rolling-win",    type=int,   default=ROLLING_WIN)
    parser.add_argument("--rolling-std",    type=float, default=ROLLING_STD)
    parser.add_argument("--plot",           action="store_true", default=SAVE_PLOTS)
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    print(f"Reading: {src}")
    df = pd.read_csv(src) if src.suffix.lower() == ".csv" else pd.read_excel(src)

    if args.datetime_col not in df.columns:
        sys.exit(f"[ERROR] Datetime column '{args.datetime_col}' not found.")

    df[args.datetime_col] = pd.to_datetime(df[args.datetime_col], infer_datetime_format=True, errors="coerce")
    df.dropna(subset=[args.datetime_col], inplace=True)
    df.set_index(args.datetime_col, inplace=True)
    df.sort_index(inplace=True)

    numeric_cols = list(df.select_dtypes(include="number").columns)
    target_cols  = args.value_cols if args.value_cols else numeric_cols
    missing      = [c for c in target_cols if c not in df.columns]
    if missing:
        sys.exit(f"[ERROR] Columns not found: {missing}")

    cfg = {
        "zscore_thresh": args.zscore_thresh,
        "iqr_factor":    args.iqr_factor,
        "rolling_win":   args.rolling_win,
        "rolling_std":   args.rolling_std,
    }

    print(f"  {len(df):,} rows | Method: {args.method} | Columns: {target_cols}\n")

    out_df = df.copy()
    summary_rows = []
    flag_col_names = []

    for col in target_cols:
        series = df[col].dropna()
        flags  = run_detection(series, args.method, cfg)
        flags  = flags.reindex(df.index, fill_value=False)

        flag_col = f"{col}_anomaly"
        out_df[flag_col] = flags.map({True: "YES", False: ""})
        flag_col_names.append(flag_col)

        count = flags.sum()
        pct   = count / len(series) * 100
        print(f"  {col}: {count} anomalies ({pct:.1f}%)")
        summary_rows.append({
            "Column":          col,
            "Total Points":    len(series),
            "Anomalies Found": count,
            "Anomaly %":       round(pct, 2),
            "Method":          args.method,
        })

        if args.plot:
            save_plot(series, flags, col, Path(args.output).parent)

    out_df.reset_index(inplace=True)

    # Determine 1-based column indices for flag columns in the output
    all_cols    = list(out_df.columns)
    flag_indices = [all_cols.index(c) + 1 for c in flag_col_names if c in all_cols]

    out = Path(args.output)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="Flagged Data", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)

    style_wb(out, {"Flagged Data": flag_indices})
    print(f"\nOutput written to: {out.resolve()}")
    print("  Red rows = anomalous data points")


if __name__ == "__main__":
    main()
  
