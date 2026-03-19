"""
ts_decompose.py
Decompose a time series into trend, seasonal, and residual components.

Dependencies: pandas, openpyxl, statsmodels, matplotlib
Install:      pip install pandas openpyxl statsmodels matplotlib

Usage:
    python ts_decompose.py --input data.csv --datetime-col "Date" --value-col "Sales"
    python ts_decompose.py --input data.csv --datetime-col "Date" --value-col "Revenue" --period 12 --model multiplicative
    python ts_decompose.py --input data.csv --datetime-col "ts" --value-col "Temp" --freq D --period 7 --plot
"""

import argparse
import sys
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from statsmodels.tsa.seasonal import seasonal_decompose

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FILE   = "data.csv"
OUTPUT_FILE  = "decomposed.xlsx"
DATETIME_COL = "Date"
VALUE_COL    = "Value"
FREQ         = None          # Resample frequency before decomposing (e.g. "MS"); None = use as-is
PERIOD       = 12            # Seasonal period: 12=monthly/annual, 7=daily/weekly, 4=quarterly/annual
MODEL        = "additive"    # additive | multiplicative
SAVE_PLOT    = True
# ─────────────────────────────────────────────────────────────────────────────

HEADER_COLOR = "1F4E79"
COLORS = {
    "observed":  "#1F4E79",
    "trend":     "#2E86AB",
    "seasonal":  "#70AD47",
    "residual":  "#FF6B35",
}


def load_series(path: Path, datetime_col: str, value_col: str,
                freq: str | None) -> pd.Series:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)

    for col in [datetime_col, value_col]:
        if col not in df.columns:
            sys.exit(f"[ERROR] Column '{col}' not found. Available: {list(df.columns)}")

    df[datetime_col] = pd.to_datetime(df[datetime_col], infer_datetime_format=True, errors="coerce")
    df.dropna(subset=[datetime_col], inplace=True)
    df.set_index(datetime_col, inplace=True)
    df.sort_index(inplace=True)

    series = pd.to_numeric(df[value_col], errors="coerce").dropna()

    if freq:
        series = series.resample(freq).mean().interpolate(method="time")

    return series


def save_decomposition_plot(result, value_col: str, model: str, out_path: Path) -> None:
    fig = plt.figure(figsize=(14, 10))
    fig.suptitle(f"Decomposition — {value_col}  ({model})", fontsize=14, fontweight="bold", y=0.98)

    gs = gridspec.GridSpec(4, 1, hspace=0.55)
    components = [
        ("observed",  result.observed,  "Observed"),
        ("trend",     result.trend,     "Trend"),
        ("seasonal",  result.seasonal,  "Seasonal"),
        ("residual",  result.resid,     "Residual"),
    ]

    for i, (key, data, label) in enumerate(components):
        ax = fig.add_subplot(gs[i])
        ax.plot(data.index, data.values, color=COLORS[key], linewidth=0.9)
        ax.set_ylabel(label, fontsize=9)
        ax.tick_params(axis="both", labelsize=8)
        ax.spines[["top", "right"]].set_visible(False)
        if key == "residual":
            ax.axhline(0, color="#aaa", linewidth=0.6, linestyle="--")

    plt.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  Chart saved: {out_path.name}")


def style_wb(path: Path) -> None:
    wb = load_workbook(path)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.freeze_panes = "A2"
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Decompose a time series into trend, seasonal, and residual.")
    parser.add_argument("--input",        default=INPUT_FILE)
    parser.add_argument("--output",       default=OUTPUT_FILE)
    parser.add_argument("--datetime-col", default=DATETIME_COL)
    parser.add_argument("--value-col",    default=VALUE_COL)
    parser.add_argument("--freq",         default=FREQ,   help="Optional resample frequency (e.g. MS, D, W)")
    parser.add_argument("--period",       type=int, default=PERIOD,
                        help="Seasonal period (12=annual monthly, 7=weekly daily, 4=annual quarterly)")
    parser.add_argument("--model",        default=MODEL, choices=["additive", "multiplicative"])
    parser.add_argument("--plot",         action="store_true", default=SAVE_PLOT)
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    print(f"Reading: {src}")
    series = load_series(src, args.datetime_col, args.value_col, args.freq)
    print(f"  {len(series):,} data points")
    print(f"  Date range: {series.index.min()} → {series.index.max()}")

    if len(series) < 2 * args.period:
        sys.exit(
            f"[ERROR] Not enough data points ({len(series)}) for period={args.period}. "
            f"Need at least {2 * args.period} points."
        )

    print(f"\nDecomposing: model={args.model}, period={args.period}")
    result = seasonal_decompose(series, model=args.model, period=args.period, extrapolate_trend="freq")

    # Assemble output DataFrame
    out_df = pd.DataFrame({
        args.datetime_col: series.index,
        "Observed":        result.observed.values,
        "Trend":           result.trend.values,
        "Seasonal":        result.seasonal.values,
        "Residual":        result.resid.values,
    })

    # Component statistics
    stats = []
    for comp, data in [("Trend", result.trend), ("Seasonal", result.seasonal), ("Residual", result.resid)]:
        stats.append({
            "Component": comp,
            "Mean":      round(data.mean(), 4),
            "Std Dev":   round(data.std(), 4),
            "Min":       round(data.min(), 4),
            "Max":       round(data.max(), 4),
        })

    # Trend direction
    import numpy as np
    x = np.arange(len(result.trend.dropna()))
    y = result.trend.dropna().values
    slope = np.polyfit(x, y, 1)[0]
    trend_direction = "Upward" if slope > 0 else "Downward"
    print(f"  Trend direction: {trend_direction} (slope={slope:.4f})")

    out = Path(args.output)
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        out_df.to_excel(writer, sheet_name="Components", index=False)
        pd.DataFrame(stats).to_excel(writer, sheet_name="Statistics", index=False)
        pd.DataFrame([{
            "Parameter": "Model",          "Value": args.model},
            {"Parameter": "Period",        "Value": args.period},
            {"Parameter": "Input points",  "Value": len(series)},
            {"Parameter": "Date range",    "Value": f"{series.index.min()} → {series.index.max()}"},
            {"Parameter": "Trend direction","Value": trend_direction},
            {"Parameter": "Trend slope",   "Value": round(slope, 6)},
        ]).to_excel(writer, sheet_name="Info", index=False)

    style_wb(out)

    if args.plot:
        plot_path = Path(args.output).with_suffix(".png")
        save_decomposition_plot(result, args.value_col, args.model, plot_path)

    print(f"\nOutput written to: {out.resolve()}")
    print("  'Components' tab — observed, trend, seasonal, residual columns")
    print("  'Statistics' tab — summary stats per component")


if __name__ == "__main__":
    main()

