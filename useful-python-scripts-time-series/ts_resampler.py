"""
ts_resampler.py
Resample and aggregate irregular time series data to a consistent frequency.

Dependencies: pandas, openpyxl
Install:      pip install pandas openpyxl

Usage:
    python ts_resampler.py --input data.csv --datetime-col "Timestamp" --freq D
    python ts_resampler.py --input data.xlsx --datetime-col "Date" --freq H --fill interpolate
    python ts_resampler.py --input data.csv  --datetime-col "ts" --freq W --config agg_rules.json

Frequency strings (--freq):
    T / min   Minutely
    H         Hourly
    D         Daily
    W         Weekly
    MS        Month start
    M         Month end
    QS        Quarter start
    AS / Y    Yearly

Example agg_rules.json:
{
    "Temperature": "mean",
    "Rainfall":    "sum",
    "WindSpeed":   "max",
    "ReadingCount":"count"
}
"""

import argparse
import json
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── CONFIG ────────────────────────────────────────────────────────────────────
INPUT_FILE   = "data.csv"
OUTPUT_FILE  = "resampled.xlsx"
DATETIME_COL = "Timestamp"
FREQ         = "D"           # Resample frequency
DEFAULT_AGG  = "mean"        # Default aggregation for all columns
FILL_METHOD  = "ffill"       # Gap filling: ffill | bfill | interpolate | none
CONFIG_FILE  = None          # Optional JSON with per-column aggregation rules
# ─────────────────────────────────────────────────────────────────────────────

HEADER_COLOR = "1F4E79"


def read_input(path: Path, datetime_col: str) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        df = pd.read_excel(path)

    if datetime_col not in df.columns:
        sys.exit(
            f"[ERROR] Datetime column '{datetime_col}' not found.\n"
            f"  Available: {list(df.columns)}"
        )

    df[datetime_col] = pd.to_datetime(df[datetime_col], infer_datetime_format=True, errors="coerce")
    n_bad = df[datetime_col].isna().sum()
    if n_bad:
        print(f"  Warning: {n_bad} rows had unparseable timestamps and will be dropped.")
    df.dropna(subset=[datetime_col], inplace=True)
    df.set_index(datetime_col, inplace=True)
    df.sort_index(inplace=True)
    return df


def build_agg_dict(df: pd.DataFrame, default_agg: str, config: dict) -> dict:
    """Build per-column aggregation mapping."""
    agg = {}
    for col in df.select_dtypes(include="number").columns:
        agg[col] = config.get(col, default_agg)
    return agg


def resample_df(df: pd.DataFrame, freq: str, agg_dict: dict,
                fill_method: str) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Resample and fill gaps. Returns (resampled_df, gap_report_df)."""
    numeric_cols = list(agg_dict.keys())
    if not numeric_cols:
        sys.exit("[ERROR] No numeric columns found to resample.")

    # Track original index to identify gaps
    original_periods = set(df[numeric_cols].resample(freq).first().index)

    resampled = df[numeric_cols].resample(freq).agg(agg_dict)
    full_index = pd.date_range(resampled.index.min(), resampled.index.max(), freq=freq)
    resampled = resampled.reindex(full_index)

    gap_periods = [str(p) for p in full_index if p not in original_periods]

    # Fill gaps
    if fill_method == "ffill":
        resampled.ffill(inplace=True)
    elif fill_method == "bfill":
        resampled.bfill(inplace=True)
    elif fill_method == "interpolate":
        resampled.interpolate(method="time", inplace=True)
    # "none" → leave NaN

    gap_report = pd.DataFrame({"Gap Period": gap_periods, "Fill Method": fill_method})
    return resampled, gap_report


def style_wb(path: Path) -> None:
    wb = load_workbook(path)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
            cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        ws.freeze_panes = "A2"
    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="Resample irregular time series data.")
    parser.add_argument("--input",        default=INPUT_FILE)
    parser.add_argument("--output",       default=OUTPUT_FILE)
    parser.add_argument("--datetime-col", default=DATETIME_COL)
    parser.add_argument("--freq",         default=FREQ, help="Resample frequency (D, H, W, M, ...)")
    parser.add_argument("--agg",          default=DEFAULT_AGG, help="Default aggregation: mean|sum|min|max|count")
    parser.add_argument("--fill",         default=FILL_METHOD, help="Gap fill: ffill|bfill|interpolate|none")
    parser.add_argument("--config",       default=CONFIG_FILE, help="JSON file with per-column aggregation rules")
    args = parser.parse_args()

    src = Path(args.input)
    if not src.exists():
        sys.exit(f"[ERROR] File not found: {src}")

    col_config = {}
    if args.config:
        cfg = Path(args.config)
        if not cfg.exists():
            sys.exit(f"[ERROR] Config file not found: {cfg}")
        with open(cfg) as fh:
            col_config = json.load(fh)

    print(f"Reading: {src}")
    df = read_input(src, args.datetime_col)
    print(f"  {len(df):,} rows after timestamp parsing")
    print(f"  Date range: {df.index.min()} → {df.index.max()}")
    print(f"  Numeric columns: {list(df.select_dtypes(include='number').columns)}")

    agg_dict = build_agg_dict(df, args.agg, col_config)
    print(f"\nResampling to frequency: {args.freq}")
    print(f"Aggregation rules: {agg_dict}")

    resampled, gap_report = resample_df(df, args.freq, agg_dict, args.fill)

    print(f"\nOutput rows : {len(resampled):,}")
    print(f"Gaps filled : {len(gap_report):,}")

    # Build summary
    summary_rows = []
    for col in resampled.columns:
        summary_rows.append({
            "Column":      col,
            "Aggregation": agg_dict.get(col, args.agg),
            "Non-null":    resampled[col].notna().sum(),
            "Null":        resampled[col].isna().sum(),
            "Min":         resampled[col].min(),
            "Max":         resampled[col].max(),
            "Mean":        round(resampled[col].mean(), 4),
        })

    out = Path(args.output)
    resampled.reset_index(inplace=True)
    resampled.rename(columns={"index": args.datetime_col}, inplace=True)

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        resampled.to_excel(writer, sheet_name="Resampled", index=False)
        pd.DataFrame(summary_rows).to_excel(writer, sheet_name="Summary", index=False)
        if not gap_report.empty:
            gap_report.to_excel(writer, sheet_name="Gap Report", index=False)

    style_wb(out)
    print(f"\nOutput written to: {out.resolve()}")


if __name__ == "__main__":
    main()

